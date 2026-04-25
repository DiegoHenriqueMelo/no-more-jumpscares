#!/usr/bin/env python3
"""Resgata logs do MongoDB e exporta para um arquivo .xlsx.

Uso:
  python scripts/exportar_logs_xlsx.py
  python scripts/exportar_logs_xlsx.py --output relatorio.xlsx
  python scripts/exportar_logs_xlsx.py --pc pc1 --sessao <id>

Variaveis de ambiente aceitas (via .env):
  MONGO_URI, MONGO_DATABASE, MONGO_COLLECTION
"""

from __future__ import annotations

import argparse
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    from pymongo import MongoClient, ASCENDING
except ImportError:
    raise SystemExit("Dependencia ausente: pymongo. Execute: pip install pymongo")

try:
    import pandas as pd
except ImportError:
    raise SystemExit("Dependencia ausente: pandas. Execute: pip install pandas openpyxl")

try:
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("Dependencia ausente: openpyxl. Execute: pip install openpyxl")


COLUNAS_ORDEM = [
    "pc",
    "sessao_treino_log",
    "noite",
    "ep",
    "resultado",
    "passos",
    "recompensa",
    "taxa_vitoria",
    "tempo_ep_minutos",
    "sessao_treino_id",
    "coletado_em_utc",
]

CABECALHOS_PT = {
    "pc": "PC",
    "sessao_treino_log": "Sessao Log",
    "noite": "Noite",
    "ep": "Episodio",
    "resultado": "Resultado",
    "passos": "Passos",
    "recompensa": "Recompensa",
    "taxa_vitoria": "Taxa Vitoria (%)",
    "tempo_ep_minutos": "Tempo EP (min)",
    "sessao_treino_id": "ID Sessao",
    "coletado_em_utc": "Coletado Em (UTC)",
}

COR_CABECALHO = "1F4E79"
COR_VITORIA   = "C6EFCE"
COR_MORTE     = "FFCCCC"
COR_LINHA_PAR = "F2F2F2"


def carregar_env(caminho: Path = Path(".env")) -> None:
    if not caminho.exists():
        return
    for linha in caminho.read_text(encoding="utf-8").splitlines():
        texto = linha.strip()
        if not texto or texto.startswith("#") or "=" not in texto:
            continue
        chave, valor = texto.split("=", 1)
        chave = chave.strip()
        valor = valor.strip().strip('"').strip("'")
        if chave and chave not in os.environ:
            os.environ[chave] = valor


def buscar_documentos(
    *,
    uri: str,
    database: str,
    collection: str,
    filtro: dict[str, Any],
) -> list[dict[str, Any]]:
    cliente = MongoClient(uri, serverSelectionTimeoutMS=8000)
    try:
        cliente.admin.command("ping")
        col = cliente[database][collection]
        cursor = col.find(filtro, {"_id": 0}).sort(
            [("pc", ASCENDING), ("sessao_treino_log", ASCENDING), ("ep", ASCENDING)]
        )
        return list(cursor)
    finally:
        cliente.close()


def documentos_para_dataframe(documentos: list[dict[str, Any]]) -> pd.DataFrame:
    df = pd.DataFrame(documentos)

    colunas_existentes = [c for c in COLUNAS_ORDEM if c in df.columns]
    extras = [c for c in df.columns if c not in COLUNAS_ORDEM]
    df = df[colunas_existentes + extras]

    df.rename(columns=CABECALHOS_PT, inplace=True)
    return df


def estilizar_planilha(ws, df: pd.DataFrame) -> None:
    fonte_cabecalho = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    fill_cabecalho  = PatternFill("solid", fgColor=COR_CABECALHO)
    fill_vitoria    = PatternFill("solid", fgColor=COR_VITORIA)
    fill_morte      = PatternFill("solid", fgColor=COR_MORTE)
    fill_par        = PatternFill("solid", fgColor=COR_LINHA_PAR)
    alinhamento_c   = Alignment(horizontal="center", vertical="center")

    col_resultado_idx = None
    for idx, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.font   = fonte_cabecalho
        cell.fill   = fill_cabecalho
        cell.alignment = alinhamento_c
        if col == CABECALHOS_PT.get("resultado", "Resultado"):
            col_resultado_idx = idx

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        resultado_val = ""
        if col_resultado_idx:
            resultado_val = str(ws.cell(row=row_idx, column=col_resultado_idx).value or "")

        if "VITORIA" in resultado_val.upper() or "WIN" in resultado_val.upper():
            fill_linha = fill_vitoria
        elif "MORTE" in resultado_val.upper() or "GAME_OVER" in resultado_val.upper():
            fill_linha = fill_morte
        elif row_idx % 2 == 0:
            fill_linha = fill_par
        else:
            fill_linha = None

        for cell in row:
            cell.alignment = Alignment(vertical="center")
            if fill_linha:
                cell.fill = fill_linha

    for col_idx, col in enumerate(df.columns, start=1):
        max_len = max(
            len(str(col)),
            df[col].astype(str).str.len().max() if not df.empty else 0,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def adicionar_aba_resumo(writer: pd.ExcelWriter, df_todos: pd.DataFrame) -> None:
    col_res      = CABECALHOS_PT.get("resultado", "Resultado")
    col_pc       = CABECALHOS_PT.get("pc", "PC")
    col_recomp   = CABECALHOS_PT.get("recompensa", "Recompensa")
    col_taxa     = CABECALHOS_PT.get("taxa_vitoria", "Taxa Vitoria (%)")
    col_ep       = CABECALHOS_PT.get("ep", "Episodio")

    if col_res not in df_todos.columns or col_pc not in df_todos.columns:
        return

    resumo = (
        df_todos.groupby(col_pc)
        .agg(
            Total_Episodios=(col_ep, "count"),
            Vitorias=(col_res, lambda x: x.str.contains("VITORIA|WIN", case=False, na=False).sum()),
            Mortes=(col_res, lambda x: x.str.contains("MORTE|GAME_OVER", case=False, na=False).sum()),
            Recompensa_Media=(col_recomp, "mean") if col_recomp in df_todos.columns else (col_ep, "count"),
            Taxa_Vitoria_Media=(col_taxa, "mean") if col_taxa in df_todos.columns else (col_ep, "count"),
        )
        .reset_index()
    )

    resumo["Taxa_Vitoria_Calculada (%)"] = (
        resumo["Vitorias"] / resumo["Total_Episodios"] * 100
    ).round(2)

    resumo.to_excel(writer, sheet_name="Resumo", index=False)

    ws = writer.sheets["Resumo"]
    fonte_cab = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    fill_cab  = PatternFill("solid", fgColor=COR_CABECALHO)
    for cell in ws[1]:
        cell.font  = fonte_cab
        cell.fill  = fill_cab
        cell.alignment = Alignment(horizontal="center")

    for col_idx, col in enumerate(resumo.columns, start=1):
        max_len = max(len(str(col)), resumo[col].astype(str).str.len().max() if not resumo.empty else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 35)

    ws.freeze_panes = "A2"


def exportar_xlsx(df: pd.DataFrame, caminho_saida: Path) -> None:
    with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
        adicionar_aba_resumo(writer, df)

        df.to_excel(writer, sheet_name="Todos os Logs", index=False)
        estilizar_planilha(writer.sheets["Todos os Logs"], df)

        col_pc = CABECALHOS_PT.get("pc", "PC")
        if col_pc in df.columns:
            for pc_id in sorted(df[col_pc].dropna().unique()):
                df_pc = df[df[col_pc] == pc_id].copy()
                nome_aba = str(pc_id)[:31]
                df_pc.to_excel(writer, sheet_name=nome_aba, index=False)
                estilizar_planilha(writer.sheets[nome_aba], df_pc)


def criar_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Exporta logs do MongoDB para .xlsx"
    )
    parser.add_argument("--output", "-o", type=Path, help="Caminho do arquivo .xlsx de saida")
    parser.add_argument("--uri", help="Mongo URI. Fallback: MONGO_URI no .env")
    parser.add_argument(
        "--database",
        default=os.getenv("MONGO_DATABASE", "no_more_jumpscares"),
    )
    parser.add_argument(
        "--collection",
        default=os.getenv("MONGO_COLLECTION", "treino_logs"),
    )
    parser.add_argument("--pc", help="Filtrar por PC (ex: pc1)")
    parser.add_argument("--sessao", dest="sessao_treino_id", help="Filtrar por sessao_treino_id")
    return parser


def main() -> None:
    carregar_env()
    parser = criar_parser()
    args = parser.parse_args()

    uri = (args.uri or os.getenv("MONGO_URI") or "").strip()
    if not uri:
        raise SystemExit("MONGO_URI ausente. Configure no .env ou passe --uri.")

    filtro: dict[str, Any] = {}
    if args.pc:
        filtro["pc"] = args.pc
    if args.sessao_treino_id:
        filtro["sessao_treino_id"] = args.sessao_treino_id

    print(f"Conectando ao MongoDB ({args.database}.{args.collection})...")
    documentos = buscar_documentos(
        uri=uri,
        database=args.database,
        collection=args.collection,
        filtro=filtro,
    )

    if not documentos:
        raise SystemExit("Nenhum documento encontrado com os filtros informados.")

    print(f"{len(documentos)} documentos encontrados.")

    df = documentos_para_dataframe(documentos)

    if args.output:
        caminho_saida = args.output
    else:
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        caminho_saida = Path(f"logs_exportados_{timestamp}.xlsx")

    exportar_xlsx(df, caminho_saida)
    print(f"Exportado com sucesso: {caminho_saida.resolve()}")


if __name__ == "__main__":
    main()
